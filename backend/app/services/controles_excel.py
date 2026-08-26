"""Exportación de los controles ESH a Excel, con el formato de las hojas en papel.

Se generan en memoria (``BytesIO``) y se devuelven con ``StreamingResponse``:
nada toca el disco del servidor, igual que el resto de las exportaciones.

La idea es que quien reciba el archivo reconozca la hoja que llenaba a mano,
así que se respetan los títulos, el orden de las columnas y las notas al pie
del formato original.
"""

import logging
from datetime import date, timedelta
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ImagenExcel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.core.constants import etiqueta_area
from app.core.controles_catalogo import (
    AREAS_PLATICAS,
    ETIQUETAS_VALOR_CHECKLIST,
    ETIQUETAS_VALOR_SI_NO,
    PUNTOS_SQP,
    RAYSER_NORMAL,
    RENGLONES_SUSTANCIAS,
    TITULO_PLATICAS,
    CampoFormato,
    DefinicionChecklist,
)
from app.models.control import InspeccionSqp
from app.services.control_service import Evidencia
# El semáforo vive en `exportacion_comun` porque también lo usa la hoja de
# estudios: así el Excel y la pantalla no se contradicen en ningún módulo.
from app.services.exportacion_comun import (
    BORDE_FINO,
    FUENTE_ENCABEZADO,
    FUENTE_TITULO,
    FUENTES_SEMAFORO,
    GRIS,
    RELLENO_ENCABEZADO,
    RELLENOS_SEMAFORO,
    ajustar_anchos,
    escribir_encabezados,
)

RELLENO_SECCION = PatternFill(start_color=GRIS, end_color=GRIS, fill_type="solid")

logger = logging.getLogger(__name__)

# Lado mayor de la evidencia dentro de la hoja, en píxeles. Una foto de celular
# a tamaño completo ocuparía varias pantallas de alto.
ANCHO_EVIDENCIA = 420

# Nombres de mes en español: `strftime("%B")` depende del locale del
# contenedor, que es C y devolvería "August".
MESES: tuple[str, ...] = (
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
)


def _dias_del_rango(desde: date, hasta: date) -> list[date]:
    """Todos los días del periodo, incluidos los extremos."""
    total = (hasta - desde).days + 1
    return [desde + timedelta(days=indice) for indice in range(total)]


def _mismo_mes(desde: date, hasta: date) -> bool:
    """``True`` si el periodo cabe en un solo mes natural."""
    return desde.year == hasta.year and desde.month == hasta.month


# --- Rayser ----------------------------------------------------------------


def _hoja_rayser(
    hoja: Worksheet, registros: list[dict[str, Any]], desde: date, hasta: date, periodo: str
) -> None:
    """Reproduce la hoja mensual de presiones."""
    hoja.title = "Rayser"

    hoja["A1"] = "CONTROL DE PRESIONES DE RAYSER"
    hoja["A1"].font = FUENTE_TITULO
    hoja.merge_cells("A1:E1")
    hoja["F1"] = f"Mes: {periodo}"
    hoja["F1"].font = Font(bold=True)

    # Con un rango dentro de un mismo mes basta el número de día, como en la
    # hoja impresa; si abarca varios, el número solo sería ambiguo.
    por_dia = _mismo_mes(desde, hasta)
    encabezados = [
        "Día" if por_dia else "Fecha",
        "Manómetro 1",
        "Manómetro 2",
        "Manómetro 3",
        "Manómetro 4",
        "Observaciones",
        "Responsable",
        "Evidencia",
    ]
    escribir_encabezados(hoja, encabezados, fila=2)

    registros_por_fecha = {registro["fecha"]: registro for registro in registros}

    fila = 3
    for dia in _dias_del_rango(desde, hasta):
        celda_dia = hoja.cell(row=fila, column=1)
        celda_dia.value = dia.day if por_dia else dia.strftime("%d/%m/%Y")
        celda_dia.alignment = Alignment(horizontal="center")
        celda_dia.border = BORDE_FINO

        registro = registros_por_fecha.get(dia)

        for indice in range(4):
            celda = hoja.cell(row=fila, column=2 + indice)
            celda.border = BORDE_FINO
            celda.alignment = Alignment(horizontal="center")

            if registro is None:
                continue

            lectura = registro["manometros"][indice]
            celda.value = float(lectura["valor"])
            celda.number_format = "0.0"
            celda.fill = RELLENOS_SEMAFORO[lectura["semaforo"]]
            celda.font = FUENTES_SEMAFORO[lectura["semaforo"]]

        celda_obs = hoja.cell(row=fila, column=6)
        celda_obs.value = registro["observaciones"] if registro else None
        celda_obs.alignment = Alignment(wrap_text=True, vertical="top")
        celda_obs.border = BORDE_FINO

        celda_resp = hoja.cell(row=fila, column=7)
        celda_resp.value = registro["responsable"] if registro else None
        celda_resp.border = BORDE_FINO

        celda_evidencia = hoja.cell(row=fila, column=8)
        if registro is not None:
            total_fotos = len(registro["fotos"])
            celda_evidencia.value = str(total_fotos) if total_fotos else "—"
        celda_evidencia.alignment = Alignment(horizontal="center")
        celda_evidencia.border = BORDE_FINO

        fila += 1

    fila += 1
    hoja.cell(
        row=fila,
        column=1,
        value=f"La presión normal de los manómetros es de {RAYSER_NORMAL:g} psi.",
    ).font = Font(italic=True)

    fila += 2
    hoja.cell(
        row=fila, column=1, value="Nombre y firma de quien realiza la inspección:"
    ).font = Font(bold=True)

    ajustar_anchos(hoja, [12, 14, 14, 14, 14, 45, 20, 12])
    hoja.freeze_panes = "A3"


def _hoja_evidencias(libro: Workbook, evidencias: list[Evidencia]) -> None:
    """Hoja con las fotos que acompañan a los registros.

    Se anclan como imágenes de la hoja, no como enlaces: el archivo se manda
    por correo y tiene que viajar completo.
    """
    hoja = libro.create_sheet("Evidencias")
    hoja["A1"] = "Evidencias fotográficas"
    hoja["A1"].font = FUENTE_TITULO
    ajustar_anchos(hoja, [70])

    fila = 3
    for evidencia in evidencias:
        titulo = f"{evidencia.fecha:%d/%m/%Y}"
        if evidencia.detalle:
            titulo += f" — {evidencia.detalle}"
        titulo += f" — registró: {evidencia.responsable}"

        encabezado = hoja.cell(row=fila, column=1, value=titulo)
        encabezado.font = Font(bold=True)

        try:
            imagen = ImagenExcel(BytesIO(evidencia.imagen))
        except Exception:
            # Un archivo ilegible (truncado al subirse, por ejemplo) no puede
            # tumbar el reporte del mes entero: se anota y se sigue.
            logger.warning(
                "Evidencia ilegible del %s; se omite en el Excel", evidencia.fecha
            )
            nota = hoja.cell(
                row=fila + 1,
                column=1,
                value=(
                    "No se pudo incrustar esta evidencia: el archivo no es una "
                    "imagen legible."
                ),
            )
            nota.font = Font(italic=True, color="9C0006")
            fila += 3
            continue

        # Se conserva la proporción: deformar la foto dificulta leer la carátula
        # del manómetro.
        if imagen.width > ANCHO_EVIDENCIA:
            proporcion = ANCHO_EVIDENCIA / float(imagen.width)
            imagen.width = ANCHO_EVIDENCIA
            imagen.height = int(imagen.height * proporcion)

        hoja.add_image(imagen, f"A{fila + 1}")

        # ~19 px por fila: se deja el alto de la imagen más un respiro.
        fila += 2 + int(imagen.height / 19) + 2


def generar_excel_rayser(
    registros: list[dict[str, Any]],
    evidencias: list[Evidencia],
    desde: date,
    hasta: date,
    periodo: str,
) -> BytesIO:
    """Arma el libro del control de presiones y lo devuelve en memoria."""
    libro = Workbook()
    hoja = libro.active
    if hoja is None:  # pragma: no cover
        hoja = libro.create_sheet()

    _hoja_rayser(hoja, registros, desde, hasta, periodo)

    if evidencias:
        _hoja_evidencias(libro, evidencias)

    flujo = BytesIO()
    libro.save(flujo)
    flujo.seek(0)
    return flujo


# --- Inspección de SQP -----------------------------------------------------


def _encabezado_sqp(hoja: Worksheet, inspeccion: InspeccionSqp) -> int:
    """Escribe la cabecera del formato y devuelve la primera fila libre."""
    hoja["A1"] = "INSPECCION DE SUSTANCIAS QUIMICAS PELIGROSAS"
    hoja["A1"].font = FUENTE_TITULO
    hoja["A1"].alignment = Alignment(horizontal="center")
    hoja.merge_cells("A1:G1")

    hoja["A2"] = "ENCARGADO Y CARGO"
    hoja["A2"].font = Font(bold=True)
    hoja.merge_cells("A2:C2")
    hoja["D2"] = "ÁREA"
    hoja["D2"].font = Font(bold=True)
    hoja.merge_cells("D2:E2")
    hoja["F2"] = "FECHA"
    hoja["F2"].font = Font(bold=True)
    hoja.merge_cells("F2:G2")

    encargado = inspeccion.encargado
    if inspeccion.cargo:
        encargado = f"{encargado} — {inspeccion.cargo}"

    hoja["A3"] = f"Nombre: {encargado}"
    hoja.merge_cells("A3:C3")
    hoja["D3"] = etiqueta_area(inspeccion.area)
    hoja.merge_cells("D3:E3")
    hoja["F3"] = inspeccion.fecha.strftime("%d/%m/%Y")
    hoja.merge_cells("F3:G3")

    return 5


def _tabla_puntos(hoja: Worksheet, inspeccion: InspeccionSqp, fila: int) -> int:
    """Escribe las secciones, los puntos y sus respuestas."""
    hoja.cell(row=fila, column=1, value="SUSTANCIAS QUÍMICAS").font = Font(bold=True)
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    _encabezados_puntos(hoja, fila)

    fila += 1
    seccion_escrita = ""
    respuestas = {respuesta.orden: respuesta for respuesta in inspeccion.respuestas}

    for orden, punto in enumerate(PUNTOS_SQP):
        if punto.seccion != seccion_escrita:
            seccion_escrita = punto.seccion
            celda_seccion = hoja.cell(row=fila, column=1, value=punto.seccion)
            celda_seccion.font = Font(bold=True)
            celda_seccion.fill = RELLENO_SECCION
            hoja.merge_cells(
                start_row=fila, start_column=1, end_row=fila, end_column=7
            )
            fila += 1

        respuesta = respuestas.get(orden)

        celda_codigo = hoja.cell(row=fila, column=1, value=punto.codigo)
        celda_codigo.alignment = Alignment(horizontal="center", vertical="top")
        celda_codigo.border = BORDE_FINO

        celda_texto = hoja.cell(row=fila, column=2, value=punto.texto)
        celda_texto.alignment = Alignment(wrap_text=True, vertical="top")
        celda_texto.border = BORDE_FINO

        # Una "X" en la columna elegida, como se marcaba a mano.
        for indice, valor in enumerate(("si", "no", "na")):
            celda = hoja.cell(row=fila, column=3 + indice)
            if respuesta is not None and respuesta.valor == valor:
                celda.value = "X"
                celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = BORDE_FINO

        celda_obs = hoja.cell(
            row=fila,
            column=6,
            value=respuesta.observaciones if respuesta else None,
        )
        celda_obs.alignment = Alignment(wrap_text=True, vertical="top")
        celda_obs.border = BORDE_FINO
        hoja.merge_cells(start_row=fila, start_column=6, end_row=fila, end_column=7)

        fila += 1

    return fila


def _celda_encabezado(hoja: Worksheet, fila: int, columna: int, texto: str) -> None:
    """Aplica el estilo de encabezado a una celda suelta.

    ``escribir_encabezados`` siempre arranca en la columna 1; aquí la fila ya
    trae contenido a la izquierda, así que se estila celda por celda.
    """
    celda = hoja.cell(row=fila, column=columna, value=texto)
    celda.fill = RELLENO_ENCABEZADO
    celda.font = FUENTE_ENCABEZADO
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    celda.border = BORDE_FINO


def _encabezados_puntos(hoja: Worksheet, fila: int) -> None:
    """Encabezados SI / NO / N/A / OBSERVACIONES de la tabla de puntos."""
    for columna, texto in ((3, "SI"), (4, "NO"), (5, "N/A"), (6, "OBSERVACIONES")):
        _celda_encabezado(hoja, fila, columna, texto)
    hoja.merge_cells(start_row=fila, start_column=6, end_row=fila, end_column=7)


def _tabla_sustancias(hoja: Worksheet, sustancias: list[str], fila: int) -> int:
    """Tabla numerada con las SQP del área."""
    fila += 1
    hoja.cell(row=fila, column=2, value="Nombre de la SQP").font = Font(bold=True)
    hoja.cell(row=fila, column=6, value="SGA").font = Font(bold=True)
    fila += 1

    # Siempre se imprimen los renglones del formato; si hay más sustancias
    # capturadas, se agregan las filas que hagan falta.
    total = max(RENGLONES_SUSTANCIAS, len(sustancias))

    for numero in range(1, total + 1):
        celda_numero = hoja.cell(row=fila, column=1, value=numero)
        celda_numero.alignment = Alignment(horizontal="center")
        celda_numero.border = BORDE_FINO

        celda_nombre = hoja.cell(
            row=fila,
            column=2,
            value=sustancias[numero - 1] if numero <= len(sustancias) else None,
        )
        celda_nombre.border = BORDE_FINO

        for columna in range(3, 8):
            hoja.cell(row=fila, column=columna).border = BORDE_FINO

        fila += 1

    return fila


def generar_excel_sqp(inspeccion: InspeccionSqp, sustancias: list[str]) -> BytesIO:
    """Arma la hoja de una inspección de SQP."""
    libro = Workbook()
    hoja = libro.active
    if hoja is None:  # pragma: no cover
        hoja = libro.create_sheet()
    hoja.title = "Inspeccion de SQP"

    fila = _encabezado_sqp(hoja, inspeccion)
    fila = _tabla_puntos(hoja, inspeccion, fila)
    fila = _tabla_sustancias(hoja, sustancias, fila)

    fila += 1
    hoja.cell(
        row=fila,
        column=1,
        value=f"Nombre de quien realiza la inspección: {inspeccion.responsable}",
    ).font = Font(bold=True)
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)

    ajustar_anchos(hoja, [8, 70, 6, 6, 7, 35, 12])

    flujo = BytesIO()
    libro.save(flujo)
    flujo.seek(0)
    return flujo


# --- Listas de verificación (OK / NO OK) -----------------------------------


def _observaciones_del_dia(registro: dict[str, Any]) -> str | None:
    """Junta las observaciones de los puntos en NO OK en una sola celda.

    La hoja en papel tiene una única columna de observaciones por día, así que
    cada texto se rotula con el punto al que pertenece.
    """
    partes = [
        f"{punto['etiqueta']}: {punto['observaciones']}"
        for punto in registro["puntos"]
        if punto["valor"] == "no_ok" and punto["observaciones"]
    ]
    return " · ".join(partes) or None


def generar_excel_checklist(
    definicion: DefinicionChecklist,
    registros: list[dict[str, Any]],
    evidencias: list[Evidencia],
    desde: date,
    hasta: date,
    periodo: str,
) -> BytesIO:
    """Arma la hoja mensual de un control de OK / NO OK.

    Sirve a los tres controles que tienen esta forma: lo único que cambia es la
    lista de columnas, que sale de la definición del catálogo.
    """
    libro = Workbook()
    hoja = libro.active
    if hoja is None:  # pragma: no cover
        hoja = libro.create_sheet()
    hoja.title = definicion.hoja

    total_columnas = len(definicion.puntos) + 3  # día, observaciones, responsable

    hoja["A1"] = definicion.titulo
    hoja["A1"].font = FUENTE_TITULO
    hoja.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=max(2, total_columnas - 1)
    )
    celda_mes = hoja.cell(row=1, column=total_columnas, value=f"Mes: {periodo}")
    celda_mes.font = Font(bold=True)

    fila_encabezados = 2
    if definicion.subtitulo:
        hoja.cell(row=2, column=1, value=definicion.subtitulo).font = Font(bold=True)
        hoja.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        fila_encabezados = 3

    por_dia = _mismo_mes(desde, hasta)
    encabezados = (
        ["Día" if por_dia else "Fecha"]
        + [punto.etiqueta for punto in definicion.puntos]
        + ["Observaciones", "Responsable"]
    )
    escribir_encabezados(hoja, encabezados, fila=fila_encabezados)

    registros_por_fecha = {registro["fecha"]: registro for registro in registros}

    fila = fila_encabezados + 1
    for dia in _dias_del_rango(desde, hasta):
        celda_dia = hoja.cell(row=fila, column=1)
        celda_dia.value = dia.day if por_dia else dia.strftime("%d/%m/%Y")
        celda_dia.alignment = Alignment(horizontal="center")
        celda_dia.border = BORDE_FINO

        registro = registros_por_fecha.get(dia)
        puntos = {punto["orden"]: punto for punto in registro["puntos"]} if registro else {}

        for indice in range(len(definicion.puntos)):
            celda = hoja.cell(row=fila, column=2 + indice)
            celda.border = BORDE_FINO
            celda.alignment = Alignment(horizontal="center")

            punto = puntos.get(indice)
            if punto is None:
                continue

            celda.value = ETIQUETAS_VALOR_CHECKLIST[punto["valor"]]
            color = "verde" if punto["valor"] == "ok" else "rojo"
            celda.fill = RELLENOS_SEMAFORO[color]
            celda.font = FUENTES_SEMAFORO[color]

        celda_obs = hoja.cell(row=fila, column=len(definicion.puntos) + 2)
        celda_obs.value = _observaciones_del_dia(registro) if registro else None
        celda_obs.alignment = Alignment(wrap_text=True, vertical="top")
        celda_obs.border = BORDE_FINO

        celda_resp = hoja.cell(row=fila, column=total_columnas)
        celda_resp.value = registro["responsable"] if registro else None
        celda_resp.border = BORDE_FINO

        fila += 1

    fila += 2
    hoja.cell(
        row=fila, column=1, value="Nombre y firma de quien realiza la inspección:"
    ).font = Font(bold=True)

    ajustar_anchos(
        hoja,
        [12] + [18] * len(definicion.puntos) + [50, 20],
    )
    hoja.freeze_panes = hoja.cell(row=fila_encabezados + 1, column=1).coordinate

    if evidencias:
        _hoja_evidencias(libro, evidencias)

    flujo = BytesIO()
    libro.save(flujo)
    flujo.seek(0)
    return flujo



# --- Formatos por inspección (silos, tableros) -----------------------------


def _texto_bilingue(espanol: str, coreano: str | None) -> str:
    """Las dos líneas del formato original, una debajo de la otra."""
    return f"{coreano}\n{espanol}" if coreano else espanol


def _fila_titulo(hoja: Worksheet, fila: int, texto: str, ancho: int) -> int:
    """Título de bloque, en gris y a lo ancho de la tabla."""
    celda = hoja.cell(row=fila, column=1, value=texto)
    celda.font = Font(bold=True)
    celda.fill = RELLENO_SECCION
    celda.alignment = Alignment(vertical="center", wrap_text=True)
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ancho)
    return fila + 1


def _bloque_campos(
    hoja: Worksheet,
    fila: int,
    campos: tuple[CampoFormato, ...],
    valores: dict[str, str],
    ancho: int,
) -> int:
    """Escribe pares etiqueta/valor, dos por renglón cuando caben."""
    columna = 1

    for campo in campos:
        etiqueta = _texto_bilingue(campo.etiqueta, campo.etiqueta_ko)
        valor = valores.get(campo.clave, "")
        if valor and campo.unidad:
            valor = f"{valor} {campo.unidad}"

        celda_etiqueta = hoja.cell(row=fila, column=columna, value=etiqueta)
        celda_etiqueta.font = Font(bold=True)
        celda_etiqueta.alignment = Alignment(vertical="top", wrap_text=True)
        celda_etiqueta.border = BORDE_FINO

        celda_valor = hoja.cell(row=fila, column=columna + 1, value=valor or "—")
        celda_valor.alignment = Alignment(vertical="top", wrap_text=True)
        celda_valor.border = BORDE_FINO

        # Los campos largos ocupan el renglón completo; el resto va de dos en
        # dos, como el encabezado de la hoja impresa.
        if campo.tipo == "texto_largo":
            hoja.merge_cells(
                start_row=fila, start_column=columna + 1, end_row=fila, end_column=ancho
            )
            fila += 1
            columna = 1
        elif columna == 1 and ancho >= 4:
            columna = 3
        else:
            fila += 1
            columna = 1

    return fila + 1 if columna != 1 else fila


def generar_excel_formato(
    definicion: DefinicionChecklist,
    registro: dict[str, Any],
    evidencias: list[Evidencia],
) -> BytesIO:
    """Reproduce la hoja de una inspección: encabezado, puntos y bloques.

    A diferencia de los controles de rejilla, aquí cada registro **es** una
    hoja: así se imprime y así se archiva.
    """
    libro = Workbook()
    hoja = libro.active
    if hoja is None:  # pragma: no cover
        hoja = libro.create_sheet()
    hoja.title = definicion.hoja[:31]

    # La tabla de puntos define el ancho de todo el documento.
    con_categoria = any(punto.categoria for punto in definicion.puntos)
    con_medicion = any(punto.medicion for punto in definicion.puntos)
    columnas = ["No."]
    if con_categoria:
        columnas.append("구분 / Categoría")
    columnas.append("점검 항목 / Punto de inspección")
    if con_medicion:
        columnas.append("측정값 / Medición")
    columnas += ["판정 / Resultado", "비고 / Observaciones"]
    ancho = len(columnas)

    celda_titulo = hoja.cell(
        row=1, column=1, value=_texto_bilingue(definicion.titulo, definicion.titulo_ko)
    )
    celda_titulo.font = FUENTE_TITULO
    celda_titulo.alignment = Alignment(vertical="center", wrap_text=True)
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    hoja.row_dimensions[1].height = 34

    fila = _bloque_campos(
        hoja,
        3,
        (CampoFormato(clave="fecha", etiqueta="Fecha", etiqueta_ko="점검일"),)
        + definicion.encabezado,
        {"fecha": f"{registro['fecha']:%d/%m/%Y}", **registro["encabezado"]},
        ancho,
    )

    fila += 1
    fila = _fila_titulo(hoja, fila, "1) 체크 리스트 / Lista de Verificación", ancho)
    escribir_encabezados(hoja, columnas, fila=fila)
    fila += 1

    etiquetas = ETIQUETAS_VALOR_SI_NO if definicion.estilo_valores == "si_no" else ETIQUETAS_VALOR_CHECKLIST

    for punto in registro["puntos"]:
        columna = 1
        celda_numero = hoja.cell(row=fila, column=columna, value=punto["orden"] + 1)
        celda_numero.alignment = Alignment(horizontal="center", vertical="top")
        celda_numero.border = BORDE_FINO
        columna += 1

        if con_categoria:
            celda_categoria = hoja.cell(
                row=fila, column=columna, value=punto.get("categoria")
            )
            celda_categoria.alignment = Alignment(vertical="top", wrap_text=True)
            celda_categoria.border = BORDE_FINO
            columna += 1

        celda_texto = hoja.cell(
            row=fila,
            column=columna,
            value=_texto_bilingue(punto["etiqueta"], punto.get("etiqueta_ko")),
        )
        celda_texto.alignment = Alignment(vertical="top", wrap_text=True)
        celda_texto.border = BORDE_FINO
        columna += 1

        if con_medicion:
            del_catalogo = definicion.puntos[punto["orden"]]
            medicion = punto.get("medicion")
            celda_medicion = hoja.cell(
                row=fila,
                column=columna,
                value=(
                    f"{medicion} {del_catalogo.medicion}"
                    if medicion and del_catalogo.medicion
                    else medicion
                ),
            )
            celda_medicion.alignment = Alignment(horizontal="center", vertical="top")
            celda_medicion.border = BORDE_FINO
            columna += 1

        celda_valor = hoja.cell(row=fila, column=columna, value=etiquetas[punto["valor"]])
        celda_valor.alignment = Alignment(horizontal="center", vertical="center")
        celda_valor.border = BORDE_FINO
        color = "verde" if punto["valor"] == "ok" else "rojo"
        celda_valor.fill = RELLENOS_SEMAFORO[color]
        celda_valor.font = FUENTES_SEMAFORO[color]
        columna += 1

        celda_obs = hoja.cell(row=fila, column=columna, value=punto.get("observaciones"))
        celda_obs.alignment = Alignment(vertical="top", wrap_text=True)
        celda_obs.border = BORDE_FINO

        fila += 1

    # --- Bloques del pie ---------------------------------------------------
    for indice, seccion in enumerate(definicion.secciones, start=2):
        valores = registro["secciones"].get(seccion.clave)
        if valores is None:
            # El bloque de anomalía no se llenó porque no hubo hallazgos.
            continue

        fila += 1
        fila = _fila_titulo(
            hoja,
            fila,
            f"{indice}) {_texto_bilingue(seccion.titulo, seccion.titulo_ko)}",
            ancho,
        )
        fila = _bloque_campos(hoja, fila, seccion.campos, valores, ancho)

    fila += 1
    hoja.cell(
        row=fila, column=1, value="점검자 서명 / Firma del inspector:"
    ).font = Font(bold=True)
    hoja.cell(
        row=fila + 1, column=1, value="확인자 서명 / Firma del supervisor:"
    ).font = Font(bold=True)

    fila += 3
    hoja.cell(
        row=fila,
        column=1,
        value=f"Capturó en el sistema: {registro['responsable']}",
    ).font = Font(italic=True)

    if definicion.nota:
        fila += 2
        celda_nota = hoja.cell(
            row=fila, column=1, value=_texto_bilingue(definicion.nota, definicion.nota_ko)
        )
        celda_nota.font = Font(bold=True, color="9C0006")
        celda_nota.alignment = Alignment(vertical="top", wrap_text=True)
        hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ancho)
        hoja.row_dimensions[fila].height = 46

    anchos = [6]
    if con_categoria:
        anchos.append(22)
    anchos.append(70)
    if con_medicion:
        anchos.append(14)
    anchos += [14, 40]
    ajustar_anchos(hoja, anchos)

    if evidencias:
        _hoja_evidencias(libro, evidencias)

    flujo = BytesIO()
    libro.save(flujo)
    flujo.seek(0)
    return flujo


# --- Pláticas diarias de seguridad -----------------------------------------


def generar_excel_platicas(
    platicas: list[dict[str, Any]],
    evidencias: list[Evidencia],
    desde: date,
    hasta: date,
    periodo: str,
) -> BytesIO:
    """Arma la hoja mensual de pláticas.

    Una columna por área, con "X" donde se impartió. Un día con varias
    pláticas ocupa varios renglones seguidos, uno por tema.
    """
    libro = Workbook()
    hoja = libro.active
    if hoja is None:  # pragma: no cover
        hoja = libro.create_sheet()
    hoja.title = "Platicas ESH"

    total_columnas = len(AREAS_PLATICAS) + 3  # día, tema, responsable

    hoja["A1"] = TITULO_PLATICAS
    hoja["A1"].font = FUENTE_TITULO
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columnas - 1)
    hoja.cell(row=1, column=total_columnas, value=f"Mes: {periodo}").font = Font(bold=True)

    por_dia = _mismo_mes(desde, hasta)
    encabezados = (
        ["Día" if por_dia else "Fecha"]
        + [area.etiqueta for area in AREAS_PLATICAS]
        + ["Tema", "Responsable"]
    )
    escribir_encabezados(hoja, encabezados, fila=2)

    por_fecha: dict[date, list[dict[str, Any]]] = {}
    for platica in platicas:
        por_fecha.setdefault(platica["fecha"], []).append(platica)

    fila = 3
    for dia in _dias_del_rango(desde, hasta):
        # Un renglón aunque no haya plática: la hoja impresa lleva los 31 días.
        # Dentro del día van en el orden en que se capturaron, no al revés como
        # en el panel: la hoja se lee de arriba abajo.
        del_dia = sorted(
            por_fecha.get(dia, []), key=lambda p: p["creado_at"]
        ) or [None]

        for platica in del_dia:
            celda_dia = hoja.cell(row=fila, column=1)
            celda_dia.value = dia.day if por_dia else dia.strftime("%d/%m/%Y")
            celda_dia.alignment = Alignment(horizontal="center")
            celda_dia.border = BORDE_FINO

            claves = (
                {area["clave"] for area in platica["areas"]} if platica else set()
            )

            for indice, area in enumerate(AREAS_PLATICAS):
                celda = hoja.cell(row=fila, column=2 + indice)
                celda.border = BORDE_FINO
                celda.alignment = Alignment(horizontal="center")

                if area.clave in claves:
                    celda.value = "X"
                    celda.font = Font(bold=True)
                    celda.fill = RELLENOS_SEMAFORO["verde"]

            celda_tema = hoja.cell(row=fila, column=len(AREAS_PLATICAS) + 2)
            celda_tema.value = platica["tema"] if platica else None
            celda_tema.alignment = Alignment(wrap_text=True, vertical="top")
            celda_tema.border = BORDE_FINO

            celda_resp = hoja.cell(row=fila, column=total_columnas)
            celda_resp.value = platica["responsable"] if platica else None
            celda_resp.border = BORDE_FINO

            fila += 1

    fila += 2
    hoja.cell(
        row=fila, column=1, value="Nombre y firma de quien realiza la inspección:"
    ).font = Font(bold=True)

    ajustar_anchos(hoja, [12] + [12] * len(AREAS_PLATICAS) + [45, 20])
    hoja.freeze_panes = "A3"

    if evidencias:
        _hoja_evidencias(libro, evidencias)

    flujo = BytesIO()
    libro.save(flujo)
    flujo.seek(0)
    return flujo

def titulo_periodo(desde: date, hasta: date) -> str:
    """Describe el periodo para el encabezado "Mes:" de la hoja."""
    if _mismo_mes(desde, hasta):
        return f"{MESES[desde.month - 1]} {desde.year}"
    return f"{desde:%d/%m/%Y} al {hasta:%d/%m/%Y}"
