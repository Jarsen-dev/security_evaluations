"""Excel del control de Extintores: tres hojas.

Va aparte de `controles_excel`, que ya pasa de mil líneas, con el mismo
criterio que `pci_excel` e `incidencias_excel`. Los estilos, la hoja de
evidencias y la cabecera de descarga sí se comparten.

**Las hojas 2 y 3 van acotadas a un periodo y la 1 no.** La ficha de los 160
aparatos es el estado de hoy y cabe entera; las revisiones son ~1 900 renglones
al día y sus evidencias son imágenes incrustadas, así que un año no cabría en
un archivo que se manda por correo.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.controles_catalogo import PUNTOS_EXTINTOR
from app.models.extintor import (
    ESTADO_CRITICO,
    ESTADO_POR_VENCER,
    ESTADO_VENCIDO,
    Extintor,
    RevisionExtintor,
    estado_vencimiento,
)
from app.services.control_service import Evidencia
from app.services.controles_excel import hoja_evidencias, titulo_periodo
from app.services.exportacion_comun import (
    FUENTE_TITULO,
    FUENTES_SEMAFORO,
    RELLENOS_SEMAFORO,
    ajustar_anchos,
    escribir_encabezados,
)

#: El color con el que se tiñe cada estado del vencimiento. Los cuatro nombres
#: son los de `RELLENOS_SEMAFORO`, que ya existen.
COLOR_ESTADO: dict[str, str] = {
    ESTADO_VENCIDO: "rojo",
    ESTADO_CRITICO: "rojo",
    ESTADO_POR_VENCER: "amarillo",
}

ETIQUETA_ESTADO: dict[str, str] = {
    ESTADO_VENCIDO: "VENCIDO",
    ESTADO_CRITICO: "Vence este mes",
    ESTADO_POR_VENCER: "Vence en dos meses",
}

ENCABEZADOS_FICHA = [
    "Folio",
    "Modelo",
    "Capacidad",
    "Tipo",
    "Ubicación",
    "Vencimiento",
    "Estado",
]
ANCHOS_FICHA = [14, 26, 14, 10, 32, 14, 22]

ANCHOS_REVISION = [12, 14, 30, 10] + [16] * len(PUNTOS_EXTINTOR) + [12, 50, 18]

VALOR_LEGIBLE = {"ok": "CONFORME", "no_ok": "INCONFORME"}


def nombre_archivo(desde: date, hasta: date) -> str:
    return f"extintores_{desde:%Y%m%d}_{hasta:%Y%m%d}.xlsx"


def _hoja_fichas(libro: Workbook, extintores: list[Extintor], hoy: date) -> None:
    """El inventario completo, con el semáforo del vencimiento."""
    hoja = libro.active
    if hoja is None:  # pragma: no cover - openpyxl siempre crea la primera
        hoja = libro.create_sheet()
    hoja.title = "Extintores"

    hoja["A1"] = "Registro de extintores"
    hoja["A1"].font = FUENTE_TITULO
    hoja["A2"] = f"Al {hoy:%d/%m/%Y} — {len(extintores)} registrados"

    escribir_encabezados(hoja, ENCABEZADOS_FICHA, fila=4)

    for extintor in extintores:
        estado = estado_vencimiento(extintor.vencimiento, hoy)
        hoja.append(
            [
                extintor.folio,
                extintor.modelo,
                extintor.capacidad,
                extintor.tipo,
                extintor.ubicacion,
                extintor.vencimiento,
                ETIQUETA_ESTADO.get(estado, "Vigente"),
            ]
        )

        color = COLOR_ESTADO.get(estado)
        if color is not None:
            # Se tiñen las dos columnas que explican el color, no la fila
            # entera: en una hoja de 160 renglones el bloque de color se lee
            # peor que el acento.
            for columna in ("F", "G"):
                celda = hoja[f"{columna}{hoja.max_row}"]
                celda.fill = RELLENOS_SEMAFORO[color]
                celda.font = FUENTES_SEMAFORO[color]

    for celda in hoja["F"][4:]:
        celda.number_format = "DD/MM/YYYY"

    ajustar_anchos(hoja, ANCHOS_FICHA)
    hoja.freeze_panes = "A5"


def _hoja_revisiones(
    libro: Workbook, revisiones: list[RevisionExtintor], desde: date, hasta: date
) -> None:
    """Una fila por revisión y los doce puntos como columnas.

    Es la forma en que se lee un control: de un vistazo se ve qué aparato falló
    y en qué punto. Una fila por punto multiplicaría por doce el alto de la
    hoja sin decir nada más.
    """
    hoja = libro.create_sheet("Revisiones")

    hoja["A1"] = "Revisiones diarias"
    hoja["A1"].font = FUENTE_TITULO
    hoja["A2"] = f"Periodo: {titulo_periodo(desde, hasta)}"

    encabezados = (
        ["Fecha", "Folio", "Ubicación", "Tipo"]
        + [punto.etiqueta for punto in PUNTOS_EXTINTOR]
        + ["Anomalías", "Observaciones", "Responsable"]
    )
    escribir_encabezados(hoja, encabezados, fila=4)

    for revision in revisiones:
        por_orden = {punto.orden: punto for punto in revision.puntos}

        observaciones = " | ".join(
            f"{PUNTOS_EXTINTOR[punto.orden].etiqueta}: {punto.observaciones}"
            for punto in revision.puntos
            if punto.valor == "no_ok" and punto.orden < len(PUNTOS_EXTINTOR)
        )

        hoja.append(
            [revision.fecha, revision.folio, revision.ubicacion, revision.tipo]
            + [
                VALOR_LEGIBLE.get(por_orden[orden].valor, "")
                if orden in por_orden
                else ""
                for orden in range(len(PUNTOS_EXTINTOR))
            ]
            + [revision.anomalias, observaciones, revision.responsable]
        )

        fila = hoja.max_row
        for orden in range(len(PUNTOS_EXTINTOR)):
            punto = por_orden.get(orden)
            if punto is None or punto.valor != "no_ok":
                continue
            # La columna 5 es el primer punto (A..D son los cuatro datos).
            celda = hoja.cell(row=fila, column=5 + orden)
            celda.fill = RELLENOS_SEMAFORO["rojo"]
            celda.font = FUENTES_SEMAFORO["rojo"]

    for celda in hoja["A"][4:]:
        celda.number_format = "DD/MM/YYYY"

    ajustar_anchos(hoja, ANCHOS_REVISION)
    hoja.freeze_panes = "E5"


def generar_excel(
    extintores: list[Extintor],
    revisiones: list[RevisionExtintor],
    evidencias: list[Evidencia],
    desde: date,
    hasta: date,
    hoy: date,
) -> BytesIO:
    """Las tres hojas del control."""
    libro = Workbook()

    _hoja_fichas(libro, extintores, hoy)
    _hoja_revisiones(libro, revisiones, desde, hasta)

    # La hoja de evidencias se crea SIEMPRE, aunque no haya fotos: el libro se
    # archiva mes con mes y una forma estable es más fácil de comparar. Mismo
    # criterio que PCI MTTO.
    if evidencias:
        hoja_evidencias(libro, evidencias)
    else:
        vacia = libro.create_sheet("Evidencias")
        vacia["A1"] = "Evidencias fotográficas"
        vacia["A1"].font = FUENTE_TITULO
        nota = vacia["A3"]
        nota.value = f"Sin evidencias registradas en {titulo_periodo(desde, hasta)}."
        nota.font = Font(italic=True, color="7F7F7F")
        ajustar_anchos(vacia, [70])

    flujo = BytesIO()
    libro.save(flujo)
    libro.close()
    flujo.seek(0)
    return flujo
