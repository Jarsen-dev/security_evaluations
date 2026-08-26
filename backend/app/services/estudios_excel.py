"""Exportación de los estudios y capacitaciones, con la maqueta de la hoja DETALLE.

Se genera en memoria (``BytesIO``) y se devuelve con ``StreamingResponse``,
igual que el resto de las exportaciones del sistema.

Las medidas —anchos de columna, gris del encabezado, color del borde, panel
congelado, autofiltro y la fila de total— están tomadas del archivo que hoy
lleva el departamento, para que quien reciba el archivo reconozca su hoja. Lo
único que se agrega es la columna ``LINK`` al final.
"""

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.estudios_catalogo import (
    APROBACIONES,
    ESTATUS,
    PRIORIDADES,
    TIPOS,
    VENCIMIENTOS,
    VIGENCIAS,
    etiqueta,
    opcion,
    semaforo,
)
from app.models.estudio import Estudio
from app.services.exportacion_comun import FUENTES_SEMAFORO, RELLENOS_SEMAFORO

NOMBRE_HOJA = "DETALLE"

# La tabla empieza en B3, con la columna A de margen: así viene el original.
FILA_ENCABEZADO = 3
PRIMERA_COLUMNA = 2

ENCABEZADOS: tuple[str, ...] = (
    "NO",
    "NO",
    "DESP",
    "ESTUDIOS",
    "한국어",
    "VIGENCIA",
    "PRIORIDAD",
    "IN/EX",
    "ESTATUS",
    "Vencido",
    "APRO",
    "PAGAR",
    "LINK",
)

# Anchos de A a N, medidos del archivo original. El de LINK es nuevo.
ANCHOS: tuple[float, ...] = (
    4.0,
    4.73,
    8.43,
    17.73,
    33.27,
    27.36,
    8.73,
    11.18,
    6.73,
    7.63,
    10.91,
    7.27,
    8.73,
    30.0,
)

# Gris del encabezado y de la fila de total, y gris más oscuro del borde.
GRIS_ENCABEZADO = "D9D9D9"
GRIS_BORDE = "A6A6A6"

RELLENO_GRIS = PatternFill(
    start_color=GRIS_ENCABEZADO, end_color=GRIS_ENCABEZADO, fill_type="solid"
)
BORDE_TABLA = Border(
    left=Side(style="thin", color=GRIS_BORDE),
    right=Side(style="thin", color=GRIS_BORDE),
    top=Side(style="thin", color=GRIS_BORDE),
    bottom=Side(style="thin", color=GRIS_BORDE),
)

CENTRADO = Alignment(horizontal="center", vertical="center")
CENTRADO_AJUSTADO = Alignment(horizontal="center", vertical="center", wrap_text=True)
# Las dos columnas de texto largo van alineadas a la izquierda y ajustadas,
# como en la hoja original.
TEXTO_LARGO = Alignment(vertical="center", wrap_text=True)

FUENTE_LINK = Font(color="0563C1", underline="single")

FORMATO_MES = "mmm-yy"

# Esquemas a los que Excel puede seguir un hipervínculo. Un texto que no
# empiece por alguno de ellos se escribe tal cual, sin ancla: el link lo teclea
# una persona y no siempre es una URL.
ESQUEMAS_ENLAZABLES: tuple[str, ...] = ("http://", "https://", "file://", "\\\\")


def _celda(
    hoja: Worksheet,
    fila: int,
    columna: int,
    valor: Any,
    *,
    alineacion: Alignment = CENTRADO,
) -> Any:
    """Escribe una celda con el borde de la tabla."""
    celda = hoja.cell(row=fila, column=columna, value=valor)
    celda.border = BORDE_TABLA
    celda.alignment = alineacion
    return celda


def _pintar(celda: Any, color: str) -> None:
    """Aplica el semáforo, si la opción tiene uno."""
    if not color:
        return
    celda.fill = RELLENOS_SEMAFORO[color]
    celda.font = FUENTES_SEMAFORO[color]


def _escribir_encabezado(hoja: Worksheet) -> None:
    """La fila de títulos, en gris y con el mismo borde que los datos."""
    for indice, texto in enumerate(ENCABEZADOS):
        celda = _celda(
            hoja,
            FILA_ENCABEZADO,
            PRIMERA_COLUMNA + indice,
            texto,
            alineacion=CENTRADO_AJUSTADO,
        )
        celda.fill = RELLENO_GRIS
        celda.font = Font(bold=True)


def _texto_vencimiento(estudio: Estudio) -> date | str:
    """La fecha cuando se conoce; si no, la etiqueta del catálogo."""
    if estudio.fecha_vencimiento is not None:
        return estudio.fecha_vencimiento
    return etiqueta(VENCIMIENTOS, estudio.vencimiento)


def _escribir_estudio(
    hoja: Worksheet, fila: int, numero: int, estudio: Estudio
) -> None:
    """Un renglón de la tabla, con sus colores."""
    _celda(hoja, fila, 2, numero)
    # La columna que la hoja original suma al pie para contar los estudios.
    _celda(hoja, fila, 3, 1)
    _celda(hoja, fila, 4, estudio.despacho)
    _celda(hoja, fila, 5, estudio.estudio, alineacion=TEXTO_LARGO)
    _celda(hoja, fila, 6, estudio.estudio_ko or "", alineacion=TEXTO_LARGO)
    _celda(hoja, fila, 7, etiqueta(VIGENCIAS, estudio.vigencia))

    # La prioridad se imprime como número (1 alta, 2 media, 3 baja), que es
    # como la lee la hoja; el color lo pone el semáforo.
    prioridad = opcion(PRIORIDADES, estudio.prioridad)
    celda = _celda(
        hoja, fila, 8, prioridad.numero if prioridad is not None else estudio.prioridad
    )
    _pintar(celda, semaforo(PRIORIDADES, estudio.prioridad))

    _celda(hoja, fila, 9, etiqueta(TIPOS, estudio.tipo))

    celda = _celda(hoja, fila, 10, etiqueta(ESTATUS, estudio.estatus).upper())
    _pintar(celda, semaforo(ESTATUS, estudio.estatus))

    celda = _celda(hoja, fila, 11, _texto_vencimiento(estudio))
    if estudio.fecha_vencimiento is not None:
        celda.number_format = FORMATO_MES

    celda = _celda(hoja, fila, 12, etiqueta(APROBACIONES, estudio.aprobado).upper())
    _pintar(celda, semaforo(APROBACIONES, estudio.aprobado))

    celda = _celda(hoja, fila, 13, etiqueta(APROBACIONES, estudio.pagado).upper())
    _pintar(celda, semaforo(APROBACIONES, estudio.pagado))

    celda = _celda(hoja, fila, 14, estudio.link or "", alineacion=TEXTO_LARGO)
    if estudio.link and estudio.link.lower().startswith(ESQUEMAS_ENLAZABLES):
        celda.hyperlink = estudio.link
        celda.font = FUENTE_LINK


def _escribir_total(hoja: Worksheet, fila: int, ultima_con_datos: int) -> None:
    """La fila gris del pie, con el conteo de estudios."""
    for columna in range(PRIMERA_COLUMNA, PRIMERA_COLUMNA + len(ENCABEZADOS)):
        celda = _celda(hoja, fila, columna, None)
        celda.fill = RELLENO_GRIS

    if ultima_con_datos >= FILA_ENCABEZADO + 1:
        hoja.cell(
            row=fila,
            column=3,
            value=f"=SUM(C{FILA_ENCABEZADO + 1}:C{ultima_con_datos})",
        )


def generar_excel_estudios(estudios: list[Estudio]) -> BytesIO:
    """Arma la hoja DETALLE con los estudios capturados."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = NOMBRE_HOJA

    # Sin cuadrícula y al 85 %, como se ve el archivo original al abrirlo.
    hoja.sheet_view.showGridLines = False
    hoja.sheet_view.zoomScale = 85

    for indice, ancho in enumerate(ANCHOS, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = ancho

    _escribir_encabezado(hoja)

    fila = FILA_ENCABEZADO
    for numero, estudio in enumerate(estudios, start=1):
        fila = FILA_ENCABEZADO + numero
        _escribir_estudio(hoja, fila, numero, estudio)

    _escribir_total(hoja, fila + 1, fila)

    # Congelar en G4 deja fijos el encabezado y las columnas de identificación
    # mientras se recorre el resto a lo ancho.
    hoja.freeze_panes = "G4"
    ultima = get_column_letter(PRIMERA_COLUMNA + len(ENCABEZADOS) - 1)
    hoja.auto_filter.ref = f"B{FILA_ENCABEZADO}:{ultima}{fila}"

    flujo = BytesIO()
    libro.save(flujo)
    flujo.seek(0)

    return flujo


def nombre_archivo_estudios() -> str:
    """``estudios_YYYYMMDD.xlsx``."""
    return f"estudios_{datetime.now().strftime('%Y%m%d')}.xlsx"
