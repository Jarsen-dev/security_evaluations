"""Excel del tablero de rondines.

Reusa el semáforo de los controles ESH para que verde y rojo signifiquen lo
mismo en todos los reportes del sistema.
"""

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.services.controles_excel import FUENTES_SEMAFORO, RELLENOS_SEMAFORO
from app.services.exportacion_comun import (
    BORDE_FINO,
    FUENTE_TITULO,
    ajustar_anchos,
    escribir_encabezados,
    slug,
)
from app.services.rondin_service import zona

NOMBRE_HOJA = "Rondines"

ETIQUETAS_TURNO = {"dia": "Día", "noche": "Noche"}

#: Celda sin visita. El guion largo se lee mejor que una celda vacía, que
#: podría confundirse con un error de generación.
SIN_VISITA = "—"

#: Bloque del turno que todavía no ocurre. Ni verde ni rojo: aún no es nada.
PENDIENTE = ""


def _local(momento: datetime) -> datetime:
    """Pasa una fecha a la hora de la planta y le quita la zona.

    No se usa `exportacion_comun.sin_zona`, que convierte a UTC: aquí el
    contenido del reporte ES la hora a la que se visitó cada punto, y verla
    corrida seis horas lo volvería inservible. openpyxl no acepta datetimes
    con zona, de ahí el `replace`.
    """
    return momento.astimezone(zona()).replace(tzinfo=None)


def nombre_reporte(fecha: date, turno: str) -> str:
    """Nombre del archivo descargado."""
    return f"rondines_{fecha:%Y%m%d}_{slug(turno)}.xlsx"


def generar_excel(tablero: dict) -> BytesIO:
    """Arma la matriz de puntos × rondines de un turno."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = NOMBRE_HOJA

    # N.º, punto, ubicación, los rondines y el total.
    total_columnas = 3 + tablero["rondines"] + 1
    transcurridos = tablero["rondines_transcurridos"]

    # --- Encabezado del reporte --------------------------------------------
    hoja.cell(row=1, column=1, value="Reporte de rondines de seguridad").font = (
        FUENTE_TITULO
    )
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columnas)

    inicio = _local(tablero["inicio"])
    fin = _local(tablero["fin"])
    hoja.cell(
        row=2,
        column=1,
        value=(
            f"Turno {ETIQUETAS_TURNO.get(tablero['turno'], tablero['turno'])}   "
            f"{inicio:%d/%m/%Y %H:%M} → {fin:%d/%m/%Y %H:%M}   "
            f"Cumplimiento: {tablero['cumplimiento']:.1f}% "
            f"({tablero['visitados']} de {tablero['total']})"
        ),
    )
    hoja.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columnas)

    # --- Tabla --------------------------------------------------------------
    encabezados = ["N.º", "Punto de control", "Ubicación"]
    encabezados += [f"Rondín {i + 1}" for i in range(tablero["rondines"])]
    encabezados.append("Visitados")

    fila_encabezado = 4
    escribir_encabezados(hoja, encabezados, fila=fila_encabezado)

    for indice, punto in enumerate(tablero["filas"]):
        fila = fila_encabezado + 1 + indice

        celda_numero = hoja.cell(row=fila, column=1, value=punto.numero)
        celda_numero.alignment = Alignment(horizontal="center")
        celda_numero.border = BORDE_FINO

        celda_nombre = hoja.cell(row=fila, column=2, value=punto.nombre)
        celda_nombre.border = BORDE_FINO

        celda_ubicacion = hoja.cell(row=fila, column=3, value=punto.ubicacion or "")
        celda_ubicacion.border = BORDE_FINO

        for columna, momento in enumerate(punto.rondines, start=4):
            celda = hoja.cell(row=fila, column=columna)
            celda.border = BORDE_FINO
            celda.alignment = Alignment(horizontal="center")

            if momento is not None:
                celda.value = f"{_local(momento):%H:%M}"
                celda.fill = RELLENOS_SEMAFORO["verde"]
                celda.font = FUENTES_SEMAFORO["verde"]
            elif columna - 4 < transcurridos:
                celda.value = SIN_VISITA
                celda.fill = RELLENOS_SEMAFORO["rojo"]
                celda.font = FUENTES_SEMAFORO["rojo"]
            else:
                # Bloque que todavía no ocurre: se deja en blanco. Pintarlo de
                # rojo lo acusaría de una falta que aún no puede haber pasado.
                celda.value = PENDIENTE

        celda_total = hoja.cell(
            row=fila,
            column=4 + tablero["rondines"],
            value=f"{punto.visitados}/{transcurridos}",
        )
        celda_total.alignment = Alignment(horizontal="center")
        celda_total.border = BORDE_FINO

    # --- Pie: porcentaje por rondín ----------------------------------------
    fila_pie = fila_encabezado + 1 + len(tablero["filas"])
    activos = tablero["puntos_activos"] or 1

    celda_pie = hoja.cell(row=fila_pie, column=2, value="Cumplimiento por rondín")
    celda_pie.border = BORDE_FINO

    for columna, visitados in enumerate(tablero["por_rondin"], start=4):
        celda = hoja.cell(row=fila_pie, column=columna)
        celda.value = f"{visitados / activos * 100:.1f}%"
        celda.alignment = Alignment(horizontal="center")
        celda.border = BORDE_FINO

    celda_general = hoja.cell(
        row=fila_pie,
        column=4 + tablero["rondines"],
        value=f"{tablero['cumplimiento']:.1f}%",
    )
    celda_general.alignment = Alignment(horizontal="center")
    celda_general.border = BORDE_FINO

    ajustar_anchos(hoja, [6, 34, 24] + [12] * tablero["rondines"] + [12])
    hoja.freeze_panes = hoja.cell(row=fila_encabezado + 1, column=4).coordinate

    # Deja legible la cabecera al imprimir en horizontal.
    hoja.print_title_rows = f"{fila_encabezado}:{fila_encabezado}"
    hoja.page_setup.orientation = "landscape"
    hoja.print_area = (
        f"A1:{get_column_letter(total_columnas)}{fila_pie}"
    )

    flujo = BytesIO()
    libro.save(flujo)
    libro.close()
    flujo.seek(0)
    return flujo
