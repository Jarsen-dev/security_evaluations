"""Exportación a Excel del control PCI MTTO.

Un libro por año, con dos hojas: los doce renglones del año y las evidencias
fotográficas. Va aparte de ``controles_excel`` por el mismo criterio que
``incidencias_excel``: aquel módulo ya pasa de mil líneas y este control no
comparte con él ni la forma de la hoja ni las columnas.

Como todas las exportaciones del proyecto, genera en ``BytesIO``: el backend no
tiene ningún volumen escribible y cualquier archivo que tocara el disco moriría
con el contenedor.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.services.control_service import Evidencia
from app.services.controles_excel import hoja_evidencias
from app.services.exportacion_comun import (
    FUENTE_TITULO,
    FUENTES_SEMAFORO,
    RELLENOS_SEMAFORO,
    ajustar_anchos,
    escribir_encabezados,
)
from app.services.pci_service import nombre_de_mes

ENCABEZADOS = (
    "Año",
    "Mes",
    "Fecha",
    "MTTO",
    "Motivo",
    "Evidencias",
    "Reporte adjunto",
    "Registró",
)

ANCHOS = [8, 14, 12, 10, 60, 12, 40, 18]


def _fila_registro(hoja, fila: int, registro: dict) -> None:
    """Escribe un renglón y lo pinta según se haya hecho el mantenimiento."""
    realizado = registro["realizado"]
    color = "verde" if realizado else "rojo"

    hoja.cell(row=fila, column=1, value=registro["anio"])
    hoja.cell(row=fila, column=2, value=nombre_de_mes(registro["mes"]).capitalize())
    hoja.cell(
        row=fila,
        column=3,
        value=f"{registro['fecha']:%d/%m/%Y}" if registro["fecha"] else "—",
    )

    # El texto ya dice SI o NO: el color refuerza, no es el único canal.
    celda = hoja.cell(row=fila, column=4, value="SI" if realizado else "NO")
    celda.fill = RELLENOS_SEMAFORO[color]
    celda.font = FUENTES_SEMAFORO[color]

    motivo = hoja.cell(row=fila, column=5, value=registro["motivo"] or "")
    if not realizado and not registro["motivo"]:
        # Una celda en blanco no distingue "no aplica" de "falta capturarlo".
        motivo.value = "Motivo pendiente"
        motivo.font = Font(italic=True, color="7F7F7F")

    hoja.cell(row=fila, column=6, value=len(registro["fotos"]))
    hoja.cell(row=fila, column=7, value=registro["reporte_nombre"] or "—")
    hoja.cell(row=fila, column=8, value=registro["responsable"])


def generar_excel_pci(
    registros: list[dict], evidencias: list[Evidencia], anio: int
) -> BytesIO:
    """El libro del año: la tabla de registros y las fotos.

    El **reporte de mantenimiento no se incrusta**: openpyxl solo sabe embeber
    imágenes, así que la hoja de registros lleva el nombre del archivo y el
    documento se baja desde el panel.
    """
    libro = Workbook()
    hoja = libro.active
    if hoja is None:  # pragma: no cover - openpyxl siempre crea la primera
        hoja = libro.create_sheet()
    hoja.title = "Registros"

    hoja["A1"] = "MANTENIMIENTO AL SISTEMA CONTRA INCENDIOS"
    hoja["A1"].font = FUENTE_TITULO
    hoja["A2"] = f"Año: {anio}"

    escribir_encabezados(hoja, list(ENCABEZADOS), fila=4)

    for indice, registro in enumerate(registros):
        _fila_registro(hoja, 5 + indice, registro)

    ajustar_anchos(hoja, ANCHOS)
    hoja.freeze_panes = hoja.cell(row=5, column=1).coordinate

    # A diferencia de los otros controles, la hoja de evidencias se crea
    # siempre: el libro se archiva año con año y una forma estable es más fácil
    # de comparar que una que aparece y desaparece.
    if evidencias:
        hoja_evidencias(libro, evidencias)
    else:
        vacia = libro.create_sheet("Evidencias")
        vacia["A1"] = "Evidencias fotográficas"
        vacia["A1"].font = FUENTE_TITULO
        nota = vacia["A3"]
        nota.value = f"Sin evidencias registradas en {anio}."
        nota.font = Font(italic=True, color="7F7F7F")
        ajustar_anchos(vacia, [70])

    flujo = BytesIO()
    libro.save(flujo)
    flujo.seek(0)
    return flujo
