"""Plantilla y lectura del Excel del catálogo de insumos.

Misma filosofía de errores que la importación de cuestionarios: **una fila mala
no invalida el archivo**. Lo estructural (archivo ilegible, hoja o columna que
falta) aborta con ``ErrorDeNegocio``; lo de cada fila se acumula y se reporta
con su número, para que quien cargó el archivo sepa exactamente qué corregir.
"""

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from app.core.constants import CATEGORIAS_INSUMO, UNIDADES_MEDIDA
from app.core.errors import ErrorDeNegocio
from app.services.exportacion_comun import ajustar_anchos, escribir_encabezados

NOMBRE_HOJA = "Insumos"
NOMBRE_HOJA_INSTRUCCIONES = "Instrucciones"

#: Tope de filas por archivo. Un catálogo de planta no llega ni cerca.
MAX_INSUMOS = 1000

COLUMNA_CODIGO = "codigo"
COLUMNA_CATEGORIA = "categoria"
COLUMNA_UNIDAD = "unidad_medida"

# El orden es el de la plantilla; el lector los localiza por nombre, así que
# reordenar las columnas en el Excel no rompe la importación.
ENCABEZADOS: list[str] = [
    "Código",
    "Descripción",
    "Categoría",
    "Unidad de medida",
    "Proveedor",
    "Ubicación",
    "Cantidad",
    "Mínimo",
    "Máximo",
]

#: Nombre de columna normalizado -> clave interna.
CLAVES: dict[str, str] = {
    "codigo": COLUMNA_CODIGO,
    "descripcion": "descripcion",
    "categoria": COLUMNA_CATEGORIA,
    "unidad de medida": COLUMNA_UNIDAD,
    "proveedor": "proveedor",
    "ubicacion": "ubicacion",
    "cantidad": "cantidad",
    "minimo": "minimo",
    "maximo": "maximo",
}

ANCHOS = [20, 42, 18, 18, 26, 24, 12, 12, 12]


@dataclass
class ErrorFila:
    """Un problema en una fila concreta del archivo."""

    fila: int
    mensaje: str


@dataclass
class ResultadoLectura:
    """Lo que se pudo leer del archivo, más lo que salió mal."""

    filas: list[dict[str, Any]] = field(default_factory=list)
    errores: list[ErrorFila] = field(default_factory=list)


def _normalizar(valor: Any) -> str:
    """Pasa una celda a texto comparable: sin acentos de más ni mayúsculas."""
    if valor is None:
        return ""
    texto = re.sub(r"\s+", " ", str(valor)).strip().casefold()
    for con, sin in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")):
        texto = texto.replace(con, sin)
    return texto


def _texto_celda(valor: Any) -> str:
    """Contenido visible de una celda, sin el ``.0`` de los floats de Excel."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return re.sub(r"\s+", " ", str(valor)).strip()


def _entero(valor: Any, campo: str) -> int:
    """Convierte una celda a entero no negativo.

    Excel entrega los números como float; el texto se acepta si representa un
    número, porque es lo que pasa cuando la columna quedó formateada como
    texto.
    """
    texto = _texto_celda(valor)
    if not texto:
        return 0

    try:
        numero = int(float(texto.replace(",", "")))
    except ValueError as exc:
        raise ValueError(f"{campo} no es un número: «{texto}».") from exc

    if numero < 0:
        raise ValueError(f"{campo} no puede ser negativo.")
    return numero


def _resolver_categoria(valor: Any) -> str:
    """Empareja el texto de la celda con una categoría del catálogo."""
    texto = _texto_celda(valor)
    if not texto:
        raise ValueError("Falta la categoría.")

    normalizada = _normalizar(texto)
    for categoria in CATEGORIAS_INSUMO:
        if _normalizar(categoria) == normalizada:
            return categoria

    raise ValueError(
        f"La categoría «{texto}» no existe. Usa una de: "
        + ", ".join(CATEGORIAS_INSUMO)
        + "."
    )


def _resolver_unidad(valor: Any) -> str:
    """Empareja el texto de la celda con una unidad de medida del catálogo."""
    texto = _texto_celda(valor)
    if not texto:
        raise ValueError("Falta la unidad de medida.")

    normalizada = _normalizar(texto)
    for unidad in UNIDADES_MEDIDA:
        if _normalizar(unidad) == normalizada:
            return unidad

    raise ValueError(
        f"La unidad «{texto}» no existe. Usa una de: "
        + ", ".join(UNIDADES_MEDIDA)
        + "."
    )


def _mapear_columnas(encabezados: tuple[Any, ...]) -> dict[str, int]:
    """Relaciona cada columna conocida con su índice en la hoja."""
    mapa: dict[str, int] = {}
    for indice, encabezado in enumerate(encabezados):
        clave = CLAVES.get(_normalizar(encabezado))
        if clave is not None and clave not in mapa:
            mapa[clave] = indice
    return mapa


def _celda(fila: tuple[Any, ...], indice: int | None) -> Any:
    """Lee una celda tolerando filas más cortas que el encabezado."""
    if indice is None or indice >= len(fila):
        return None
    return fila[indice]


def parsear_excel(contenido: bytes) -> ResultadoLectura:
    """Lee el archivo y devuelve las filas aprovechables y los errores."""
    try:
        # read_only evita cargar la hoja completa en memoria; data_only trae
        # el resultado de las fórmulas en vez de la fórmula misma.
        libro = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl lanza varios tipos según el daño
        raise ErrorDeNegocio(
            "No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx)."
        ) from exc

    try:
        hoja = None
        for nombre in libro.sheetnames:
            if _normalizar(nombre) == _normalizar(NOMBRE_HOJA):
                hoja = libro[nombre]
                break

        if hoja is None:
            raise ErrorDeNegocio(
                f"El archivo debe tener una hoja llamada '{NOMBRE_HOJA}'. "
                f"Descarga la plantilla para ver el formato esperado."
            )

        filas = hoja.iter_rows(values_only=True)

        try:
            encabezados = next(filas)
        except StopIteration as exc:
            raise ErrorDeNegocio("El archivo está vacío.") from exc

        mapa = _mapear_columnas(encabezados)
        for obligatoria, etiqueta in (
            (COLUMNA_CODIGO, "Código"),
            (COLUMNA_CATEGORIA, "Categoría"),
            (COLUMNA_UNIDAD, "Unidad de medida"),
        ):
            if obligatoria not in mapa:
                raise ErrorDeNegocio(
                    f"Falta la columna '{etiqueta}'. Descarga la plantilla "
                    f"para ver el formato esperado."
                )

        resultado = ResultadoLectura()

        # La fila 1 es el encabezado, así que la numeración arranca en 2 y
        # coincide con lo que ve el usuario en Excel.
        for numero, fila in enumerate(filas, start=2):
            codigo = _texto_celda(_celda(fila, mapa.get(COLUMNA_CODIGO)))
            if not codigo:
                # Fila sin código: separador visual, se salta en silencio.
                continue

            if len(resultado.filas) >= MAX_INSUMOS:
                resultado.errores.append(
                    ErrorFila(
                        numero,
                        f"Se alcanzó el límite de {MAX_INSUMOS} insumos por "
                        f"archivo; el resto no se leyó.",
                    )
                )
                break

            try:
                datos = {
                    "codigo": codigo,
                    "descripcion": _texto_celda(_celda(fila, mapa.get("descripcion")))
                    or None,
                    "categoria": _resolver_categoria(
                        _celda(fila, mapa.get(COLUMNA_CATEGORIA))
                    ),
                    "unidad_medida": _resolver_unidad(
                        _celda(fila, mapa.get(COLUMNA_UNIDAD))
                    ),
                    "proveedor": _texto_celda(_celda(fila, mapa.get("proveedor")))
                    or None,
                    "ubicacion": _texto_celda(_celda(fila, mapa.get("ubicacion")))
                    or None,
                    "cantidad": _entero(_celda(fila, mapa.get("cantidad")), "Cantidad"),
                    "minimo": _entero(_celda(fila, mapa.get("minimo")), "El mínimo"),
                    "maximo": _entero(_celda(fila, mapa.get("maximo")), "El máximo"),
                }
            except ValueError as exc:
                resultado.errores.append(ErrorFila(numero, str(exc)))
                continue

            if datos["maximo"] < datos["minimo"]:
                resultado.errores.append(
                    ErrorFila(
                        numero,
                        "El máximo de inventario no puede ser menor que el mínimo.",
                    )
                )
                continue

            resultado.filas.append(datos)

        return resultado
    finally:
        # En modo read_only openpyxl deja abierto el archivo temporal.
        libro.close()


def generar_plantilla() -> BytesIO:
    """Arma el Excel de ejemplo que se descarga desde el panel."""
    libro = Workbook()

    hoja = libro.active
    hoja.title = NOMBRE_HOJA
    escribir_encabezados(hoja, ENCABEZADOS)
    hoja.append(
        [
            "GN-100-M",
            "Guantes de nitrilo talla M, caja con 100 piezas",
            "EPP",
            "PZA",
            "Suministros Industriales del Norte",
            "Almacén — anaquel A3",
            120,
            50,
            200,
        ]
    )
    ajustar_anchos(hoja, ANCHOS)
    hoja.freeze_panes = "A2"

    instrucciones = libro.create_sheet(NOMBRE_HOJA_INSTRUCCIONES)
    lineas: list[tuple[str, bool]] = [
        ("Cómo llenar esta plantilla", True),
        ("", False),
        ("Captura un insumo por renglón en la hoja «Insumos».", False),
        ("Borra el renglón de ejemplo antes de importar.", False),
        ("", False),
        ("Columnas obligatorias", True),
        ("Código: identifica al insumo. No puede repetirse.", False),
        ("Categoría: una de las de abajo, tal cual está escrita.", False),
        ("Unidad de medida: una de las de abajo, tal cual está escrita.", False),
        ("", False),
        ("Columnas opcionales", True),
        ("Descripción, Proveedor y Ubicación pueden ir vacías.", False),
        ("Cantidad, Mínimo y Máximo: números enteros. Vacío cuenta como 0.", False),
        ("El máximo no puede ser menor que el mínimo.", False),
        ("", False),
        ("Categorías válidas", True),
        *[(f"• {categoria}", False) for categoria in CATEGORIAS_INSUMO],
        ("", False),
        ("Unidades de medida válidas", True),
        *[(f"• {unidad}", False) for unidad in UNIDADES_MEDIDA],
        ("", False),
        ("Qué pasa al importar", True),
        ("Los insumos nuevos se dan de alta.", False),
        ("Los que ya existen se omiten: el archivo no pisa lo capturado.", False),
        ("Una fila con problemas no invalida el resto; se reporta su número.", False),
    ]

    for texto, es_titulo in lineas:
        celda = instrucciones.cell(row=instrucciones.max_row + 1, column=1, value=texto)
        if es_titulo:
            celda.font = Font(bold=True, size=12)
        celda.alignment = Alignment(vertical="center")

    ajustar_anchos(instrucciones, [90])

    flujo = BytesIO()
    libro.save(flujo)
    libro.close()
    flujo.seek(0)
    return flujo
