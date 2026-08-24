"""Utilidades compartidas por las exportaciones a Excel y PowerPoint."""

import re
import unicodedata
from urllib.parse import quote
from dataclasses import dataclass
from datetime import UTC, datetime
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.cuestionario import Cuestionario
from app.services import estadistica_service
from app.services.estadistica_service import Filtros

TOP_PREGUNTAS = 10
# Un área por debajo de este porcentaje de su meta se marca como rezagada en
# las conclusiones de la presentación.
UMBRAL_PARTICIPACION = 80


@dataclass
class DatosReporte:
    """Todo lo que necesitan las dos exportaciones, consultado una sola vez."""

    cuestionario: Cuestionario
    resumen: dict[str, Any]
    por_area: list[dict[str, Any]]
    por_pregunta: list[dict[str, Any]]
    distribucion: list[dict[str, Any]]
    linea_tiempo: list[dict[str, Any]]
    columnas_preguntas: list[dict[str, Any]]
    filas_intentos: list[dict[str, Any]]
    generado_at: datetime


async def reunir_datos(
    db: AsyncSession, cuestionario: Cuestionario, filtros: Filtros
) -> DatosReporte:
    """Ejecuta todas las agregaciones necesarias para un reporte."""
    columnas, filas = await estadistica_service.detalle_intentos(db, filtros)

    return DatosReporte(
        cuestionario=cuestionario,
        resumen=await estadistica_service.resumen(db, filtros),
        por_area=await estadistica_service.por_area(db, filtros),
        por_pregunta=await estadistica_service.por_pregunta(db, filtros),
        distribucion=await estadistica_service.distribucion(db, filtros),
        linea_tiempo=await estadistica_service.linea_tiempo(db, filtros),
        columnas_preguntas=columnas,
        filas_intentos=filas,
        generado_at=datetime.now(UTC),
    )


# --- Estilos compartidos de las hojas de Excel -----------------------------
#
# Los usan tanto el reporte de evaluaciones (``excel_export``) como los
# formatos de los controles ESH (``controles_excel``), para que todo lo que
# salga del sistema se vea como un solo documento.

AZUL = "1F4E79"
GRIS = "F2F2F2"

FUENTE_ENCABEZADO = Font(bold=True, color="FFFFFF", size=11)
RELLENO_ENCABEZADO = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
FUENTE_TITULO = Font(bold=True, size=14)
BORDE_FINO = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

FORMATO_FECHA = "dd/mm/yyyy hh:mm"
FORMATO_PORCENTAJE = "0.00"


def escribir_encabezados(
    hoja: Worksheet, encabezados: list[str], fila: int = 1
) -> None:
    """Escribe una fila de encabezados con el estilo del sistema."""
    for columna, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=fila, column=columna, value=texto)
        celda.fill = RELLENO_ENCABEZADO
        celda.font = FUENTE_ENCABEZADO
        celda.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        celda.border = BORDE_FINO


def ajustar_anchos(hoja: Worksheet, anchos: list[int]) -> None:
    """Fija el ancho de cada columna, de izquierda a derecha."""
    for indice, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = ancho


def slug(texto: str) -> str:
    """Convierte el nombre del cuestionario en algo seguro para un archivo.

    Quita acentos y deja solo letras, números y guiones bajos: un nombre con
    "ñ" o "/" rompería la cabecera Content-Disposition o el sistema de
    archivos de quien lo descargue.
    """
    sin_acentos = (
        unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    )
    limpio = re.sub(r"[^a-zA-Z0-9]+", "_", sin_acentos).strip("_").lower()
    return limpio[:60] or "cuestionario"


def nombre_archivo(cuestionario: Cuestionario, extension: str) -> str:
    """Arma `evaluacion_{nombre_slug}_{YYYYMMDD}.{ext}`."""
    fecha = datetime.now(UTC).strftime("%Y%m%d")
    return f"evaluacion_{slug(cuestionario.nombre)}_{fecha}.{extension}"


def sin_zona(momento: datetime | None) -> datetime | None:
    """Quita la zona horaria de una fecha.

    Las columnas son TIMESTAMPTZ, pero Excel no soporta datetimes con zona y
    openpyxl lanza una excepción al escribirlos. Se convierte a UTC y se
    descarta el tzinfo.
    """
    if momento is None:
        return None
    return momento.astimezone(UTC).replace(tzinfo=None)


def formatear_duracion(segundos: int | None) -> str:
    """Segundos a `mm:ss`, o guion si el intento no se finalizó."""
    if segundos is None:
        return "—"
    return f"{segundos // 60}:{segundos % 60:02d}"


def periodo_texto(filtros: Filtros) -> str:
    """Describe el rango de fechas para la portada y la hoja de resumen."""
    if filtros.desde and filtros.hasta:
        return f"Del {filtros.desde:%d/%m/%Y} al {filtros.hasta:%d/%m/%Y}"
    if filtros.desde:
        return f"Desde el {filtros.desde:%d/%m/%Y}"
    if filtros.hasta:
        return f"Hasta el {filtros.hasta:%d/%m/%Y}"
    return "Todo el periodo"


def preguntas_mas_falladas(
    por_pregunta: list[dict[str, Any]], limite: int = TOP_PREGUNTAS
) -> list[dict[str, Any]]:
    """Ordena las preguntas de mayor a menor índice de error."""
    con_datos = [p for p in por_pregunta if p["total_respuestas"] > 0]
    return sorted(con_datos, key=lambda p: p["porcentaje_error"] or 0, reverse=True)[
        :limite
    ]


def generar_conclusiones(datos: DatosReporte) -> list[str]:
    """Arma los bullets de la última diapositiva a partir de reglas fijas.

    No son texto libre: salen de comparar cada área contra el promedio
    general y contra su meta, más las tres preguntas peor contestadas.
    """
    conclusiones: list[str] = []
    promedio_general = datos.resumen["promedio_general"]

    if promedio_general is None:
        return ["Aún no hay respuestas suficientes para generar conclusiones."]

    con_intentos = [area for area in datos.por_area if area["intentos"] > 0]

    bajo_promedio = [
        area
        for area in con_intentos
        if area["promedio"] is not None and area["promedio"] < promedio_general
    ]
    if bajo_promedio:
        nombres = ", ".join(
            f"{area['label']} ({area['promedio']:.1f}%)"
            for area in sorted(bajo_promedio, key=lambda a: a["promedio"] or 0)
        )
        conclusiones.append(
            f"Áreas por debajo del promedio general ({promedio_general:.1f}%): {nombres}."
        )
    else:
        conclusiones.append(
            f"Todas las áreas alcanzaron o superaron el promedio general "
            f"({promedio_general:.1f}%)."
        )

    rezagadas = [
        area
        for area in datos.por_area
        if area["porcentaje_participacion"] is not None
        and area["porcentaje_participacion"] < UMBRAL_PARTICIPACION
    ]
    if rezagadas:
        nombres = ", ".join(
            f"{area['label']} ({area['porcentaje_participacion']:.0f}%)"
            for area in sorted(rezagadas, key=lambda a: a["porcentaje_participacion"] or 0)
        )
        conclusiones.append(
            f"Áreas con participación menor al {UMBRAL_PARTICIPACION}% de su meta: {nombres}."
        )
    else:
        conclusiones.append(
            f"Todas las áreas con meta capturada superaron el {UMBRAL_PARTICIPACION}% "
            f"de participación."
        )

    peores = preguntas_mas_falladas(datos.por_pregunta, limite=3)
    if peores:
        conclusiones.append("Temas que requieren recapacitación prioritaria:")
        conclusiones.extend(
            f"    {indice}. {pregunta['texto']} — {pregunta['porcentaje_error']:.0f}% de error"
            for indice, pregunta in enumerate(peores, start=1)
        )

    sin_meta = [area["label"] for area in datos.por_area if area["meta"] is None]
    if sin_meta:
        conclusiones.append(
            f"Sin meta de headcount capturada (no se calculó su participación): "
            f"{', '.join(sin_meta)}."
        )

    return conclusiones


def cabecera_descarga(nombre: str) -> dict[str, str]:
    """Arma el Content-Disposition de una descarga.

    Los nombres se generan sin acentos, pero se agrega la variante
    ``filename*`` por si alguno llegara a incluir caracteres fuera de ASCII:
    sin ella, algunos navegadores truncan el nombre del archivo.
    """
    return {
        "Content-Disposition": (
            f'attachment; filename="{nombre}"; filename*=UTF-8\'\'{quote(nombre)}'
        )
    }
