"""Exportación del reporte a Excel: cuatro hojas, generado en memoria."""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.exportacion_comun import (
    FORMATO_FECHA,
    FORMATO_PORCENTAJE,
    GRIS,
    DatosReporte,
    ajustar_anchos,
    escribir_encabezados,
    formatear_duracion,
    sin_zona,
)


def _hoja_resumen(libro: Workbook, datos: DatosReporte, periodo: str) -> None:
    """Hoja 1: identificación del reporte y KPIs generales."""
    hoja = libro.active
    if hoja is None:  # pragma: no cover
        hoja = libro.create_sheet()
    hoja.title = "Resumen"

    hoja["A1"] = "Reporte de evaluación de conocimientos"
    hoja["A1"].font = Font(bold=True, size=16)

    resumen = datos.resumen
    participacion = resumen["participacion"]

    filas: list[tuple[str, Any]] = [
        ("Cuestionario", datos.cuestionario.nombre),
        ("Descripción", datos.cuestionario.descripcion or "—"),
        ("Periodo", periodo),
        ("Fecha de generación", datos.generado_at.strftime("%d/%m/%Y %H:%M UTC")),
        ("", ""),
        ("Respuestas recibidas", resumen["total_respuestas"]),
        ("Intentos sin finalizar", resumen["total_en_progreso"]),
        (
            "Meta de participación",
            participacion["meta"] if participacion["meta"] else "Sin capturar",
        ),
        (
            "Nivel de participación",
            f"{participacion['porcentaje']:.2f}%"
            if participacion["porcentaje"] is not None
            else "Sin meta capturada",
        ),
        (
            "Calificación promedio",
            f"{resumen['promedio_general']:.2f}%"
            if resumen["promedio_general"] is not None
            else "—",
        ),
        ("Aprobados", resumen["aprobados"]),
        (
            "Tasa de aprobación",
            f"{resumen['tasa_aprobacion']:.2f}%"
            if resumen["tasa_aprobacion"] is not None
            else "—",
        ),
        ("Umbral de aprobación usado", f"{resumen['umbral_aprobacion']}%"),
    ]

    for indice, (etiqueta, valor) in enumerate(filas, start=3):
        celda_etiqueta = hoja.cell(row=indice, column=1, value=etiqueta)
        celda_etiqueta.font = Font(bold=True)
        celda_etiqueta.fill = PatternFill(start_color=GRIS, end_color=GRIS, fill_type="solid")
        hoja.cell(row=indice, column=2, value=valor)

    ajustar_anchos(hoja, [30, 55])


def _hoja_detalle(libro: Workbook, datos: DatosReporte) -> None:
    """Hoja 2: un renglón por intento y una columna por pregunta."""
    hoja = libro.create_sheet("Respuestas detalladas")

    encabezados = [
        "Nombre",
        "Núm. empleado",
        "Área",
        "Fecha inicio",
        "Fecha fin",
        "Duración",
        "Correctas",
        "Total",
        "Puntaje %",
    ]
    encabezados += [
        f"P{columna['orden'] + 1}. {columna['texto']}"
        for columna in datos.columnas_preguntas
    ]

    escribir_encabezados(hoja, encabezados)

    for numero, fila in enumerate(datos.filas_intentos, start=2):
        hoja.cell(row=numero, column=1, value=fila["nombre"])
        hoja.cell(row=numero, column=2, value=fila["numero_empleado"])
        hoja.cell(row=numero, column=3, value=fila["area"])

        celda_inicio = hoja.cell(row=numero, column=4, value=sin_zona(fila["iniciado_at"]))
        celda_inicio.number_format = FORMATO_FECHA

        if fila["finalizado_at"] is not None:
            celda_fin = hoja.cell(row=numero, column=5, value=sin_zona(fila["finalizado_at"]))
            celda_fin.number_format = FORMATO_FECHA

        hoja.cell(row=numero, column=6, value=formatear_duracion(fila["duracion_segundos"]))
        hoja.cell(row=numero, column=7, value=fila["correctas"])
        hoja.cell(row=numero, column=8, value=fila["total_preguntas"])

        celda_puntaje = hoja.cell(row=numero, column=9, value=fila["puntaje"])
        celda_puntaje.number_format = FORMATO_PORCENTAJE

        for indice, columna in enumerate(datos.columnas_preguntas):
            elegida = fila["respuestas"].get(columna["id"])
            celda = hoja.cell(row=numero, column=10 + indice)
            if elegida is None:
                celda.value = "Sin responder"
                celda.font = Font(color="999999", italic=True)
            else:
                texto, es_correcta = elegida
                celda.value = texto
                # Verde para acierto, rojo para error: permite leer patrones
                # de un vistazo sin tener que cruzar con la hoja de preguntas.
                celda.font = Font(color="0F7B2F" if es_correcta else "C02626")

    # Congelar encabezado y primera columna: con 8 preguntas y 200 filas es
    # imposible navegar sin esto.
    hoja.freeze_panes = "B2"
    hoja.auto_filter.ref = (
        f"A1:{get_column_letter(len(encabezados))}{max(2, len(datos.filas_intentos) + 1)}"
    )

    ajustar_anchos(hoja, [24, 15, 16, 18, 18, 10, 10, 8, 11])
    for indice in range(len(datos.columnas_preguntas)):
        hoja.column_dimensions[get_column_letter(10 + indice)].width = 28


def _hoja_por_area(libro: Workbook, datos: DatosReporte) -> None:
    """Hoja 3: agregados por área."""
    hoja = libro.create_sheet("Por área")

    escribir_encabezados(
        hoja,
        [
            "Área",
            "Intentos",
            "Promedio %",
            "Mínimo %",
            "Máximo %",
            "Aprobados",
            "% Aprobación",
            "Meta",
            "% Participación",
        ],
    )

    for numero, area in enumerate(datos.por_area, start=2):
        hoja.cell(row=numero, column=1, value=area["label"])
        hoja.cell(row=numero, column=2, value=area["intentos"])

        for columna, clave in enumerate(
            ["promedio", "minimo", "maximo"], start=3
        ):
            celda = hoja.cell(row=numero, column=columna, value=area[clave])
            celda.number_format = FORMATO_PORCENTAJE

        hoja.cell(row=numero, column=6, value=area["aprobados"])

        celda_aprobacion = hoja.cell(
            row=numero, column=7, value=area["porcentaje_aprobacion"]
        )
        celda_aprobacion.number_format = FORMATO_PORCENTAJE

        # Sin meta capturada se escribe el texto en vez de un 0 engañoso.
        hoja.cell(row=numero, column=8, value=area["meta"] if area["meta"] else "Sin capturar")

        if area["porcentaje_participacion"] is not None:
            celda_participacion = hoja.cell(
                row=numero, column=9, value=area["porcentaje_participacion"]
            )
            celda_participacion.number_format = FORMATO_PORCENTAJE
        else:
            hoja.cell(row=numero, column=9, value="—")

    hoja.freeze_panes = "A2"
    ajustar_anchos(hoja, [20, 10, 13, 11, 11, 12, 14, 10, 16])


def _hoja_por_pregunta(libro: Workbook, datos: DatosReporte) -> None:
    """Hoja 4: acierto y error por pregunta, con desglose de opciones."""
    hoja = libro.create_sheet("Por pregunta")

    max_opciones = max(
        (len(pregunta["opciones"]) for pregunta in datos.por_pregunta), default=0
    )

    encabezados = ["#", "Pregunta", "Respuestas", "Correctas", "% Acierto", "% Error"]
    encabezados += [f"Opción {indice + 1}" for indice in range(max_opciones)]

    escribir_encabezados(hoja, encabezados)

    for numero, pregunta in enumerate(datos.por_pregunta, start=2):
        hoja.cell(row=numero, column=1, value=pregunta["orden"] + 1)

        celda_texto = hoja.cell(row=numero, column=2, value=pregunta["texto"])
        celda_texto.alignment = Alignment(wrap_text=True, vertical="top")

        hoja.cell(row=numero, column=3, value=pregunta["total_respuestas"])
        hoja.cell(row=numero, column=4, value=pregunta["correctas"])

        celda_acierto = hoja.cell(row=numero, column=5, value=pregunta["porcentaje_acierto"])
        celda_acierto.number_format = FORMATO_PORCENTAJE

        celda_error = hoja.cell(row=numero, column=6, value=pregunta["porcentaje_error"])
        celda_error.number_format = FORMATO_PORCENTAJE

        for indice, opcion in enumerate(pregunta["opciones"]):
            marca = "✔ " if opcion["es_correcta"] else ""
            celda = hoja.cell(
                row=numero,
                column=7 + indice,
                value=f"{marca}{opcion['texto']}: {opcion['elegida']} ({opcion['porcentaje']:.1f}%)",
            )
            if opcion["es_correcta"]:
                celda.font = Font(bold=True, color="0F7B2F")

    # Escala de color sobre el % de acierto: rojo abajo, verde arriba.
    if datos.por_pregunta:
        ultima = len(datos.por_pregunta) + 1
        hoja.conditional_formatting.add(
            f"E2:E{ultima}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=50,
                mid_color="FFEB84",
                end_type="num",
                end_value=100,
                end_color="63BE7B",
            ),
        )

    hoja.freeze_panes = "C2"
    ajustar_anchos(hoja, [5, 50, 12, 11, 11, 11])
    for indice in range(max_opciones):
        hoja.column_dimensions[get_column_letter(7 + indice)].width = 30


def generar_excel(datos: DatosReporte, periodo: str) -> BytesIO:
    """Arma el libro de cuatro hojas y lo devuelve en memoria."""
    libro = Workbook()

    _hoja_resumen(libro, datos, periodo)
    _hoja_detalle(libro, datos)
    _hoja_por_area(libro, datos)
    _hoja_por_pregunta(libro, datos)

    flujo = BytesIO()
    libro.save(flujo)
    libro.close()
    flujo.seek(0)

    return flujo
