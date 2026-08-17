"""Importación de preguntas desde un archivo Excel.

Filosofía de errores: una fila mala no invalida el archivo completo. Se
procesa todo lo que se pueda y se devuelve un reporte con el número de fila
de cada problema, para que el usuario corrija su Excel y vuelva a intentar.
"""

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.errors import ErrorDeNegocio
from app.schemas.cuestionario import OpcionIn, PreguntaIn

NOMBRE_HOJA = "Preguntas"
NOMBRE_HOJA_INSTRUCCIONES = "Instrucciones"

MAX_PREGUNTAS = 200
MAX_OPCIONES = 5
MIN_OPCIONES = 2

COLUMNA_PREGUNTA = "pregunta"
COLUMNA_RESPUESTA = "respuesta correcta"

ENCABEZADOS = [
    "Pregunta",
    "Opcion 1",
    "Opcion 2",
    "Opcion 3",
    "Opcion 4",
    "Opcion 5",
    "Respuesta Correcta",
]


@dataclass
class ErrorFila:
    """Problema detectado en una fila concreta del archivo."""

    fila: int
    mensaje: str


@dataclass
class ResultadoImportacion:
    """Reporte de la importación."""

    importadas: int = 0
    errores: list[ErrorFila] = field(default_factory=list)
    # Preguntas listas para agregarse al constructor del frontend.
    preguntas: list[PreguntaIn] = field(default_factory=list)


def _normalizar(valor: Any) -> str:
    """Pasa una celda a texto comparable: sin espacios sobrantes y en minúsculas."""
    if valor is None:
        return ""
    return re.sub(r"\s+", " ", str(valor)).strip().casefold()


def _texto_celda(valor: Any) -> str:
    """Devuelve el contenido visible de una celda, sin espacios sobrantes.

    Los números se convierten sin el ``.0`` que arrastran los floats de
    Excel: una opción "2020" no debe quedar como "2020.0".
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return re.sub(r"\s+", " ", str(valor)).strip()


def _mapear_columnas(encabezados: tuple[Any, ...]) -> dict[str, int]:
    """Relaciona cada nombre de columna con su índice.

    Se acepta "Opcion 1" y "Opción 1": el usuario escribe el encabezado a
    mano con frecuencia y el acento no debe romper la importación.
    """
    mapa: dict[str, int] = {}

    for indice, encabezado in enumerate(encabezados):
        nombre = _normalizar(encabezado).replace("ó", "o")

        if nombre == COLUMNA_PREGUNTA:
            mapa[COLUMNA_PREGUNTA] = indice
        elif nombre == COLUMNA_RESPUESTA:
            mapa[COLUMNA_RESPUESTA] = indice
        else:
            coincidencia = re.fullmatch(r"opcion\s*(\d+)", nombre)
            if coincidencia:
                numero = int(coincidencia.group(1))
                if 1 <= numero <= MAX_OPCIONES:
                    mapa[f"opcion {numero}"] = indice

    return mapa


def _resolver_correcta(
    valor_correcta: Any, opciones: list[str]
) -> tuple[int | None, str | None]:
    """Determina qué opción es la correcta.

    Acepta el número de opción (1–5) o el texto exacto, normalizando espacios
    y mayúsculas. Devuelve ``(indice, None)`` o ``(None, mensaje_de_error)``.
    """
    crudo = _texto_celda(valor_correcta)

    if crudo == "":
        return None, "Falta la respuesta correcta."

    # Por número de opción. Tiene prioridad, como indica la plantilla.
    if re.fullmatch(r"\d+", crudo):
        numero = int(crudo)
        if 1 <= numero <= len(opciones):
            return numero - 1, None
        # Fuera de rango no es necesariamente un error: cuando las opciones
        # son números (años, códigos de parte), lo que el usuario escribió es
        # el texto de la opción. Se sigue al match por texto.

    # Por texto exacto de la opción.
    normalizada = _normalizar(crudo)
    coincidencias = [
        indice
        for indice, texto in enumerate(opciones)
        if _normalizar(texto) == normalizada
    ]

    if len(coincidencias) == 1:
        return coincidencias[0], None

    if len(coincidencias) > 1:
        return None, (
            f"La respuesta correcta '{crudo}' coincide con más de una opción; "
            f"usa el número de opción para evitar la ambigüedad."
        )

    return None, f"La respuesta correcta '{crudo}' no corresponde a ninguna opción."


def parsear_excel(contenido: bytes) -> ResultadoImportacion:
    """Lee el archivo y devuelve las preguntas válidas junto con los errores."""
    try:
        # read_only evita cargar la hoja completa en memoria; data_only trae
        # el resultado de las fórmulas en vez de la fórmula misma.
        libro = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl lanza varios tipos según el daño
        raise ErrorDeNegocio(
            "No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx)."
        ) from exc

    try:
        hoja = None
        for nombre in libro.sheetnames:
            if _normalizar(nombre) == _normalizar(NOMBRE_HOJA):
                hoja = libro[nombre]
                break

        if hoja is None:
            raise ErrorDeNegocio(
                f"El archivo debe tener una hoja llamada '{NOMBRE_HOJA}'. "
                f"Descarga la plantilla para ver el formato esperado."
            )

        filas = hoja.iter_rows(values_only=True)

        try:
            encabezados = next(filas)
        except StopIteration as exc:
            raise ErrorDeNegocio("El archivo está vacío.") from exc

        columnas = _mapear_columnas(encabezados)

        if COLUMNA_PREGUNTA not in columnas:
            raise ErrorDeNegocio(
                "No se encontró la columna 'Pregunta'. Descarga la plantilla "
                "para ver el formato esperado."
            )
        if COLUMNA_RESPUESTA not in columnas:
            raise ErrorDeNegocio(
                "No se encontró la columna 'Respuesta Correcta'. Descarga la "
                "plantilla para ver el formato esperado."
            )

        indices_opciones = [
            columnas[f"opcion {numero}"]
            for numero in range(1, MAX_OPCIONES + 1)
            if f"opcion {numero}" in columnas
        ]

        if not indices_opciones:
            raise ErrorDeNegocio(
                "No se encontró ninguna columna de opciones ('Opcion 1', "
                "'Opcion 2', …)."
            )

        resultado = ResultadoImportacion()

        # La fila 1 es el encabezado, así que el contenido arranca en la 2.
        for numero_fila, fila in enumerate(filas, start=2):
            texto_pregunta = _texto_celda(_celda(fila, columnas[COLUMNA_PREGUNTA]))

            # Fila separadora: se salta en silencio, según la especificación.
            if texto_pregunta == "":
                continue

            if len(resultado.preguntas) >= MAX_PREGUNTAS:
                resultado.errores.append(
                    ErrorFila(
                        fila=numero_fila,
                        mensaje=(
                            f"El archivo excede el máximo de {MAX_PREGUNTAS} "
                            f"preguntas; esta fila y las siguientes se omitieron."
                        ),
                    )
                )
                break

            opciones = [
                texto
                for indice in indices_opciones
                if (texto := _texto_celda(_celda(fila, indice))) != ""
            ]

            if len(opciones) < MIN_OPCIONES:
                resultado.errores.append(
                    ErrorFila(
                        fila=numero_fila,
                        mensaje=(
                            f"Solo tiene {len(opciones)} opción(es); se requieren "
                            f"mínimo {MIN_OPCIONES}."
                        ),
                    )
                )
                continue

            indice_correcta, error = _resolver_correcta(
                _celda(fila, columnas[COLUMNA_RESPUESTA]), opciones
            )

            if error is not None or indice_correcta is None:
                resultado.errores.append(
                    ErrorFila(fila=numero_fila, mensaje=error or "Fila inválida.")
                )
                continue

            resultado.preguntas.append(
                PreguntaIn(
                    texto=texto_pregunta,
                    puntos=1,
                    opciones=[
                        OpcionIn(texto=texto, es_correcta=(indice == indice_correcta))
                        for indice, texto in enumerate(opciones)
                    ],
                )
            )

        resultado.importadas = len(resultado.preguntas)
        return resultado
    finally:
        # En modo read_only openpyxl deja abierto el archivo temporal.
        libro.close()


def _celda(fila: tuple[Any, ...], indice: int) -> Any:
    """Lee una celda tolerando filas más cortas que el encabezado."""
    return fila[indice] if indice < len(fila) else None


def generar_plantilla() -> BytesIO:
    """Crea la plantilla vacía con encabezados, ejemplo e instrucciones."""
    libro = Workbook()

    hoja = libro.active
    if hoja is None:  # pragma: no cover - Workbook() siempre crea una hoja
        raise ErrorDeNegocio("No se pudo generar la plantilla.")
    hoja.title = NOMBRE_HOJA

    relleno = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fuente_encabezado = Font(bold=True, color="FFFFFF")

    for columna, encabezado in enumerate(ENCABEZADOS, start=1):
        celda = hoja.cell(row=1, column=columna, value=encabezado)
        celda.fill = relleno
        celda.font = fuente_encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center")

    ejemplo = [
        "¿Cuál es el EPP obligatorio en el área de moldes?",
        "Casco",
        "Guantes térmicos",
        "Botas dieléctricas",
        "Todas las anteriores",
        "",
        4,
    ]
    for columna, valor in enumerate(ejemplo, start=1):
        hoja.cell(row=2, column=columna, value=valor)

    anchos = [55, 22, 22, 22, 22, 22, 20]
    for columna, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    # Congelar el encabezado: con 200 preguntas es indispensable.
    hoja.freeze_panes = "A2"

    instrucciones = libro.create_sheet(NOMBRE_HOJA_INSTRUCCIONES)
    lineas = [
        ("Cómo llenar esta plantilla", True),
        ("", False),
        ("1. Escribe una pregunta por fila en la hoja 'Preguntas'.", False),
        ("2. Las columnas 'Opcion 1' y 'Opcion 2' son obligatorias.", False),
        ("3. Las columnas 'Opcion 3', 'Opcion 4' y 'Opcion 5' son opcionales;", False),
        ("   si las dejas vacías simplemente se ignoran.", False),
        ("", False),
        ("4. En 'Respuesta Correcta' puedes escribir:", False),
        ("   - El NÚMERO de la opción correcta (1, 2, 3, 4 o 5), o", False),
        ("   - El TEXTO exacto de esa opción.", False),
        ("   No importan los espacios de más ni las mayúsculas.", False),
        ("", False),
        ("5. Las filas con la columna 'Pregunta' vacía se saltan:", False),
        ("   puedes usarlas como separadores.", False),
        ("", False),
        (f"6. Máximo {MAX_PREGUNTAS} preguntas por archivo.", False),
        ("", False),
        ("7. No cambies el nombre de la hoja ni el de las columnas.", False),
        ("", False),
        ("Si una fila tiene un error, el sistema importa el resto y te muestra", False),
        ("el número de fila con el problema para que lo corrijas.", False),
        ("", False),
        ("La fila de ejemplo de la hoja 'Preguntas' puede borrarse.", False),
    ]

    for numero, (texto, es_titulo) in enumerate(lineas, start=1):
        celda = instrucciones.cell(row=numero, column=1, value=texto)
        if es_titulo:
            celda.font = Font(bold=True, size=14)

    instrucciones.column_dimensions["A"].width = 85

    flujo = BytesIO()
    libro.save(flujo)
    libro.close()
    flujo.seek(0)

    return flujo
