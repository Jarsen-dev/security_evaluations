"""Excel de la pestaña de Incidencias.

Un renglón por hoja con hallazgos, con todo el recorrido del problema: qué se
detectó, quién lo levantó, qué se hizo, quién lo resolvió y a qué hora quedó
cerrado. Las fotos —las del hallazgo y las de la verificación— van en la hoja
de evidencias que ya arma ``controles_excel``, para que el archivo viaje
completo por correo sin depender de ninguna liga.

Se genera en ``BytesIO`` y se devuelve con ``StreamingResponse``: nada toca el
disco del servidor, que además no tiene ningún volumen escribible.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.core.controles_catalogo import CONTROLES_CHECKLIST
from app.services.cierre_service import (
    CONTROL_RAYSER,
    CONTROL_SQP,
    IncidenciaCompleta,
)
from app.services.controles_excel import hoja_evidencias
from app.services.exportacion_comun import (
    BORDE_FINO,
    FUENTES_SEMAFORO,
    RELLENOS_SEMAFORO,
    ajustar_anchos,
    escribir_encabezados,
)

#: Nombre legible de cada control. Los de lista de verificación salen del
#: catálogo; los otros dos no están ahí porque no son checklists.
NOMBRES_CONTROL: dict[str, str] = {
    CONTROL_RAYSER: "Rayser",
    CONTROL_SQP: "Inspección de SQP",
}

ENCABEZADOS = (
    "Fecha",
    "Control",
    "Identificación",
    "Problemas",
    "Descripción del hallazgo",
    "Estado",
    "Hora de hallazgo",
    "Ubicación",
    "Acción inmediata realizada",
    "Departamento o responsable",
    "Hora de cierre",
    "Acción pendiente",
    "Levantó",
    "Cerró",
)

ANCHOS = [12, 24, 18, 10, 50, 14, 14, 26, 50, 26, 12, 40, 16, 16]

#: Columna del semáforo, para no buscarla por nombre en cada renglón.
COLUMNA_ESTADO: int = ENCABEZADOS.index("Estado") + 1


def _nombre_control(clave: str) -> str:
    """Cómo se llama el control en la hoja."""
    definicion = CONTROLES_CHECKLIST.get(clave)
    if definicion is not None:
        return definicion.hoja
    return NOMBRES_CONTROL.get(clave, clave)


def _descripcion(completa: IncidenciaCompleta) -> str:
    """Los problemas de la hoja, uno por renglón dentro de la celda.

    Se arma con la etiqueta del punto y sus observaciones, que es lo mismo que
    el panel muestra como "descripción" del cierre: el texto no se guarda
    duplicado en ningún lado.
    """
    lineas: list[str] = []

    for hallazgo in completa.hallazgos:
        if hallazgo.observaciones:
            lineas.append(f"{hallazgo.etiqueta}: {hallazgo.observaciones}")
        else:
            lineas.append(hallazgo.etiqueta)

    return "\n".join(lineas)


def generar_excel_incidencias(
    incidencias: list[IncidenciaCompleta],
    titulo_periodo: str,
) -> BytesIO:
    """Arma el libro con las incidencias del periodo y sus evidencias."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Incidencias"

    hoja["A1"] = "Incidencias de los controles ESH"
    hoja["A1"].font = Font(bold=True, size=14)
    hoja["A2"] = titulo_periodo

    escribir_encabezados(hoja, ENCABEZADOS, fila=4)
    ajustar_anchos(hoja, ANCHOS)

    fila = 5
    for completa in incidencias:
        incidencia = completa.incidencia
        cierre = incidencia.cierre

        # Un cierre con acción pendiente está resuelto pero arrastra algo: se
        # pinta distinto para que no se confunda con uno limpio.
        if cierre is None:
            estado, color = "Pendiente", "rojo"
        elif cierre.accion_pendiente:
            estado, color = "Cerrado con pendiente", "naranja"
        else:
            estado, color = "Cerrado", "verde"

        valores = (
            incidencia.fecha,
            _nombre_control(incidencia.control),
            incidencia.identificacion or "—",
            incidencia.total_hallazgos,
            _descripcion(completa),
            estado,
            cierre.hora_hallazgo if cierre else "",
            cierre.ubicacion if cierre else "",
            cierre.accion_inmediata if cierre else "",
            cierre.responsable_accion if cierre else "",
            cierre.hora_cierre if cierre else "",
            (cierre.accion_pendiente or "") if cierre else "",
            incidencia.responsable,
            cierre.responsable if cierre else "",
        )

        for columna, valor in enumerate(valores, start=1):
            celda = hoja.cell(row=fila, column=columna, value=valor)
            celda.border = BORDE_FINO
            celda.alignment = Alignment(wrap_text=True, vertical="top")

        celda_estado = hoja.cell(row=fila, column=COLUMNA_ESTADO)
        celda_estado.fill = RELLENOS_SEMAFORO[color]
        celda_estado.font = FUENTES_SEMAFORO[color]

        hoja.cell(row=fila, column=1).number_format = "DD/MM/YYYY"
        fila += 1

    hoja.freeze_panes = hoja.cell(row=5, column=1).coordinate

    # Las fotos del hallazgo y las de la verificación, en la misma hoja.
    evidencias = [
        evidencia for completa in incidencias for evidencia in completa.evidencias
    ]
    if evidencias:
        hoja_evidencias(libro, evidencias)

    flujo = BytesIO()
    libro.save(flujo)
    flujo.seek(0)

    return flujo
